import pyodbc
import pandas as pd
import openpyxl

def opc(all):
    df=openpyxl.load_workbook(r'opc.xlsx')

    df2=df.active
    op=[]

    for row in range(0,df2.max_row):
        op.append((df2.cell(row+1,1).value,df2.cell(row+1,2).value,df2.cell(row+1,3).value,df2.cell(row+1,4).value
                ,df2.cell(row+1,5).value,df2.cell(row+1,6).value))

    print(op)
    opc=[]
    L2=""
    L3=""
    L4=""
    L5=""
    for i in all:
        if len(i)==20:
            for j in op:

                if str(j[0]) == str(i[18:20]):
                    opc.append((i,j[1],j[2],j[3],j[4],j[5]))

        if len(i)==18:
            for j in op:

                if str(j[0]) == "A":
                    opc.append((i,j[1],j[2],j[3],j[4],j[5]))



    
    print(opc)            

    f=pd.DataFrame(opc,columns=op[0])
    return(f)