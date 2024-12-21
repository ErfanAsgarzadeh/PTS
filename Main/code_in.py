import pyodbc
import pandas as pd
import openpyxl



def codeIn(CodeCurr):
    df=openpyxl.load_workbook(r'Code_in.xlsx')

    df2=df.active
    code=[]

    for row in range(0,df2.max_row):
        code.append((df2.cell(row+1,1).value,df2.cell(row+1,2).value,df2.cell(row+1,3).value,
                df2.cell(row+1,4).value,df2.cell(row+1,5).value,df2.cell(row+1,6).value,
                df2.cell(row+1,7).value,df2.cell(row+1,8).value,df2.cell(row+1,9).value,
                df2.cell(row+1,10).value,df2.cell(row+1,11).value,df2.cell(row+1,12).value,
                df2.cell(row+1,13).value))

    print(code)
    codeIn=[]
    for i in CodeCurr:
        print(i)
        if len(i[0])==20:
            for j in code:
                print((i[0])[18:20])
                if str(j[0]) == str(i[0][18:20]):
                    codeIn.append((i[0],i[1],j[2],j[3],j[4],(i[0])[0:2],j[6],j[7],j[8],j[9],j[10],j[11],j[12]))
        if len(i[0])==18:
            for j in code:
                print((i[0])[18:20])
                if str(j[0]) == "A":
                    codeIn.append((i[0],i[1],j[2],j[3],j[4],(i[0])[0:2],j[6],j[7],j[8],j[9],j[10],j[11],j[12]))
        
        if len(i[0])==13:
            for j in code:
                if str(j[0]) == "Z9":
                    codeIn.append((i[0],i[1],j[2],j[3],j[4],(i[0])[0:2],j[6],j[7],j[8],j[9],j[10],j[11],j[12]))
    print(codeIn)

    codeInp=list(dict.fromkeys(codeIn))            

    f=pd.DataFrame(codeInp,columns=code[0])
    return(f)