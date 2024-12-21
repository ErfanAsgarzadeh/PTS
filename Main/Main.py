import sqlite3
from tkinter import filedialog

import openpyxl
import pandas as pd
import pyodbc

import bom
import code_in
import opc
import qcp

file_path = filedialog.askopenfilename()
df=openpyxl.load_workbook(file_path)
df1=df.active
all=[]
for row in range(0,df1.max_row):
    for col in df1.iter_cols(1,df1.max_column):
        if col[row].value != None :
            all.append(col[row].value)
print(len(all))
df.close()

try:
    conn = pyodbc.connect("Driver={ODBC Driver 17 for SQL Server};"
                      r"Server=192.168.10.230\MSSQLSERVER2022 ;"
                      "Database=MISPTS;"
                      "UID=sa;PWD=Almas123!;Trusted_Connection=no")
except:
    print("gdgd")


mcu=conn.cursor()

NewCode=[]
oldcode=[]
for code in all:
    mcu.execute("SELECT [PrdCode] FROM vw_InvPrdDef where PrdCode='"+code+"'")
    data=mcu.fetchall()
    try:
        A=data[0][0]
        oldcode.append(code)
    except:
        NewCode.append(code)
        continue
        #print("Duplicated")    

print(NewCode)
f=pd.DataFrame(oldcode)
f.to_excel("oldcode.xlsx")


conn = sqlite3.connect('ee.db')


b=pd.Series

def exchange_code (code):
    exc = conn.cursor()
    a="'"+code[0]+code[1]+"'"
    B="SELECT ord,PARA_NAME,STRU_DESC FROM STRUCTURE WHERE STRU_CODE="+a
    exc.execute(B)
    b=exc.fetchall()
    exc.close

    b.sort()
    #print(b)
    

    descrip=""
    n1=2
    try:
        descrip=b[0][2]
    except:
        print("not defined")



    for i in b:
        exc.execute("SELECT CODE FROM PARAMETER WHERE PARA_NAME="+"'"+i[1]+"'")
        a=(exc.fetchall())
        exc.close
        for n in a:
            de_no=len(n[0])
        para_code=code[n1:n1+de_no]
        n1=n1+de_no
        #print(para_code)
        exc.execute("SELECT DESC FROM PARAMETER WHERE PARA_NAME="+"'"+i[1]+"' and CODE="+"'"+para_code+"'")
        
        gh=(exc.fetchall())
        exc.close
        for c in gh:
            descrip=descrip+(c[0])+" "
    
    #print(descrip)
    return (code,descrip)

ex_code=[]
N=1
M=1
i=1
CodeCurr=NewCode
for code in CodeCurr:
    ex_code.append( exchange_code(code))
    #print(ex_code)
    print(N,":",len(CodeCurr))
    N=N+1



df=pd.DataFrame(ex_code)
df.to_excel("Exchanged.xlsx")

f=code_in.codeIn(ex_code)
f.to_excel("codeList.xlsx")

f=opc.opc(all)
f.to_excel("opcList.xlsx")

f=bom.bom(all)
f.to_excel("bomList.xlsx")

f=qcp.qcp(all)
f.to_excel("qcpList.xlsx")