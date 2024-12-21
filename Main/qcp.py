import pyodbc
import pandas as pd
import openpyxl


def qcp(all):
    df=openpyxl.load_workbook(r'qcp.xlsx')

    df2=df.active
    qc=[]

    for row in range(0,df2.max_row):
        qc.append((df2.cell(row+1,1).value,df2.cell(row+1,2).value))

    print(qc)
    qcp=[]
    for i in all:
        if len(i)==20:
            for j in qc:
                #print(i[18:20])
                if str(j[0]) == str(i[18:20]):
                    qcp.append((i,j[1]))

    print(qcp)
    qcpDup=list(dict.fromkeys(qcp))            

    f=pd.DataFrame(qcpDup)
    return f