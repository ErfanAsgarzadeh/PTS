import pyodbc
import pandas as pd
import openpyxl


def bom(all):
    df=openpyxl.load_workbook(r'bom.xlsx')

    df2=df.active
    bo=[]

    for row in range(0,df2.max_row):
        bo.append((df2.cell(row+1,1).value,df2.cell(row+1,2).value,df2.cell(row+1,3).value,df2.cell(row+1,4).value
                ,df2.cell(row+1,5).value,df2.cell(row+1,6).value,df2.cell(row+1,7).value,df2.cell(row+1,8).value
                ,df2.cell(row+1,9).value,df2.cell(row+1,10).value,df2.cell(row+1,11).value,df2.cell(row+1,12).value
                ,df2.cell(row+1,13).value,df2.cell(row+1,14).value))

    print(bo)
    bom=[]
    L2=""
    L3=""
    L4=""
    L5=""
    for i in all:
        if len(i)==20:
            for j in bo:
                #print(i[18:20])
                if str(j[0]) == str(i[18:20]):

                    if str(j[5])!="None":
                        L2=str(i[:18])+str(j[5])
                        L3=str(j[6])
                    elif str(j[5])=="None":
                        L2=""
                        L3=""

                    if str(j[11])!="None":
                        L4=str(i[:18])+str(j[11])
                        L5=str(j[12])
                    elif str(j[11])=="None":
                        L4=""
                        L5=""    


                    bom.append((i,j[1],j[2],j[3],j[4],L2,L3,j[7],j[8],j[9],j[10],L4, L5,j[13]))

        if len(i)==18:
            for j in bo:
                #print(i[18:20])
                if str(j[0]) == "A":
                    if str(j[5])=="00":
                        if i[0]=="A":
                            L2="Y"+str(i[1:17])+"200"
                        else:
                            L2="R"+str(i[1:17])+"200"
                        bom.append((i,j[1],j[2],j[3],j[4],L2,L3,j[7],j[8],j[9],j[10],"", "",j[13]))
                    if str(j[5])=="NP":
                        L2="F117000003"
                        bom.append((i,j[1],j[2],j[3],j[4],L2,L3,j[7],j[8],j[9],j[10],"", "",j[13]))


    
    print(bom)            

    f=pd.DataFrame(bom,columns=bo[0])
    return f
